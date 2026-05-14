"""Data export routes."""

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
import csv
import json
import io
from datetime import datetime
from backend.analytics import AnalyticsService
from backend.services import WorkspaceService

try:
    from openpyxl import Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

router = APIRouter()

workspace_service = WorkspaceService()


@router.get("/csv")
async def export_csv():
    """Export all workspace data as CSV."""
    try:
        db_path = workspace_service.get_database_path()
        analytics = AnalyticsService(db_path)
        
        # Get all data
        emp_data = analytics.get_employee_workload()
        proj_data = analytics.get_project_health()
        task_data = analytics.get_task_distribution()
        risk_data = analytics.get_risk_analysis()
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Export employee data
        writer.writerow(['=== EMPLOYEE DATA ==='])
        if 'employees' in emp_data:
            writer.writerow(['Employee ID', 'Name', 'Total Tasks', 'Completed', 'Overdue', 'Completion Rate'])
            for emp in emp_data['employees']:
                writer.writerow([
                    emp.get('employee_id', ''),
                    emp.get('name', ''),
                    emp.get('total_tasks', 0),
                    emp.get('completed_tasks', 0),
                    emp.get('overdue_tasks', 0),
                    emp.get('completion_rate', 0)
                ])
        
        writer.writerow([])
        writer.writerow(['=== PROJECT DATA ==='])
        if 'projects' in proj_data:
            writer.writerow(['Project ID', 'Name', 'Status', 'Health Score', 'Tasks'])
            for proj in proj_data['projects']:
                writer.writerow([
                    proj.get('project_id', ''),
                    proj.get('name', ''),
                    proj.get('status', ''),
                    proj.get('health_score', 0),
                    proj.get('task_count', 0)
                ])
        
        writer.writerow([])
        writer.writerow(['=== TASK DATA ==='])
        writer.writerow(['Status', 'Count'])
        if 'status_distribution' in task_data:
            for status, count in task_data['status_distribution'].items():
                writer.writerow([status, count])
        
        writer.writerow([])
        writer.writerow(['=== RISK DATA ==='])
        writer.writerow(['Risk Level', 'Count'])
        if 'risk_distribution' in risk_data:
            for risk, count in risk_data['risk_distribution'].items():
                writer.writerow([risk, count])
        
        # Convert to bytes
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=workflow_data.csv"}
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting CSV: {str(e)}")


@router.get("/json")
async def export_json():
    """Export analytics data as JSON."""
    try:
        db_path = workspace_service.get_database_path()
        analytics = AnalyticsService(db_path)
        
        # Collect all analytics data
        data = {
            "employees": analytics.get_employee_workload(),
            "projects": analytics.get_project_health(),
            "tasks": analytics.get_task_distribution(),
            "productivity": analytics.get_employee_productivity(),
            "risks": analytics.get_risk_analysis(),
            "export_timestamp": datetime.utcnow().isoformat()
        }
        
        json_str = json.dumps(data, indent=2, default=str)
        
        return StreamingResponse(
            iter([json_str]),
            media_type="application/json",
            headers={"Content-Disposition": "attachment; filename=workflow_data.json"}
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting JSON: {str(e)}")


@router.get("/excel")
async def export_excel():
    """Export analytics data as Excel file."""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=400, detail="Excel export not available. Install openpyxl: pip install openpyxl")
    
    try:
        db_path = workspace_service.get_database_path()
        analytics = AnalyticsService(db_path)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        # Employee data sheet
        try:
            emp_data = analytics.get_employee_workload()
            if emp_data and 'employees' in emp_data and emp_data['employees']:
                ws = wb.create_sheet("Employees")
                ws.append(['Employee ID', 'Name', 'Total Tasks', 'Completed', 'Overdue', 'Completion Rate'])
                for emp in emp_data['employees']:
                    ws.append([
                        emp.get('id', ''),
                        emp.get('name', ''),
                        emp.get('total_tasks', 0),
                        emp.get('completed_tasks', 0),
                        emp.get('overdue_tasks', 0),
                        round(emp.get('completion_rate', 0), 2)
                    ])
        except Exception as e:
            pass
        
        # Project data sheet
        try:
            proj_data = analytics.get_project_health()
            if proj_data and 'projects' in proj_data and proj_data['projects']:
                ws = wb.create_sheet("Projects")
                ws.append(['Project ID', 'Name', 'Status', 'Health Score', 'Tasks'])
                for proj in proj_data['projects']:
                    ws.append([
                        proj.get('id', ''),
                        proj.get('name', ''),
                        proj.get('status', ''),
                        round(proj.get('health_score', 0), 2),
                        proj.get('task_count', 0)
                    ])
        except Exception as e:
            pass
        
        # Task data sheet
        try:
            task_data = analytics.get_task_distribution()
            if task_data:
                ws = wb.create_sheet("Tasks")
                ws.append(['Status', 'Count'])
                if 'status_distribution' in task_data:
                    for status, count in task_data['status_distribution'].items():
                        ws.append([status, count])
                
                ws.append([])
                ws.append(['Priority', 'Count'])
                if 'priority_distribution' in task_data:
                    for priority, count in task_data['priority_distribution'].items():
                        ws.append([priority, count])
        except Exception as e:
            pass
        
        # Productivity sheet
        try:
            prod_data = analytics.get_employee_productivity()
            if prod_data:
                ws = wb.create_sheet("Productivity")
                ws.append(['Metric', 'Value'])
                if 'metrics' in prod_data:
                    for key, value in prod_data['metrics'].items():
                        ws.append([key, value])
                elif 'productivity' in prod_data:
                    for key, value in prod_data['productivity'].items():
                        ws.append([key, value])
        except Exception as e:
            pass
        
        # Risk data sheet
        try:
            risk_data = analytics.get_risk_analysis()
            if risk_data:
                ws = wb.create_sheet("Risks")
                ws.append(['Risk Level', 'Count'])
                if 'risk_distribution' in risk_data:
                    for risk, count in risk_data['risk_distribution'].items():
                        ws.append([risk, count])
        except Exception as e:
            pass
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=workflow_data.xlsx"}
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting Excel: {str(e)}")
